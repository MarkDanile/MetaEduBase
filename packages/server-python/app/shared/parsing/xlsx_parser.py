"""Excel row extraction using openpyxl."""

from __future__ import annotations

from dataclasses import dataclass, field


@dataclass
class ParsedDataset:
    column_names: list[str] = field(default_factory=list)
    column_types: list[str] = field(default_factory=list)
    rows: list[dict[str, str]] = field(default_factory=list)


def extract_xlsx_rows(file_path: str) -> ParsedDataset:
    """Extract rows from an Excel file. First row is treated as headers."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    first_row = next(rows_iter, None)

    if first_row is None:
        wb.close()
        return ParsedDataset()

    column_names = [str(cell or f"col_{i}") for i, cell in enumerate(first_row)]
    column_types: list[str] = []
    data_rows: list[dict[str, str]] = []

    second_row = next(rows_iter, None)
    if second_row:
        column_types = [_infer_type(cell) for cell in second_row]
        row_dict = {column_names[i]: str(cell or "") for i, cell in enumerate(second_row) if i < len(column_names)}
        data_rows.append(row_dict)

    for row in rows_iter:
        row_dict = {column_names[i]: str(cell or "") for i, cell in enumerate(row) if i < len(column_names)}
        data_rows.append(row_dict)

    wb.close()
    return ParsedDataset(
        column_names=column_names,
        column_types=column_types,
        rows=data_rows,
    )


def _infer_type(value: object) -> str:
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "float"
    from datetime import date, datetime
    if isinstance(value, (date, datetime)):
        return "date"
    return "string"
