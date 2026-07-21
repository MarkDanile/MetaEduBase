"""Tests for the review-required cooperation-notes generator."""
from __future__ import annotations

from openpyxl import Workbook, load_workbook

from tests.scripts._script_loader import load_server_script

generate = load_server_script("generate_cooperation_notes").generate


def test_generate_marks_every_row_synthetic_and_review_required(tmp_path):
    source = tmp_path / "05_客户.xlsx"
    output = tmp_path / "13_客户_合作跟进记录_待审核.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["客户ID", "所属行业", "来源渠道"])
    sheet.append(["C-1", "人工智能", "政府引荐"])
    workbook.save(source)

    count = generate(source, output, limit=10)

    assert count == 1
    generated = load_workbook(output, read_only=False, data_only=True)
    row = next(generated.active.iter_rows(min_row=2, values_only=True))
    headers = [cell.value for cell in generated.active[1]]
    values = dict(zip(headers, row, strict=True))
    generated.close()
    assert values["客户ID"] == "C-1"
    assert values["数据来源"] == "synthetic"
    assert values["审核状态"] == "待审核"
    assert "非真实客户沟通事实" in values["备注"]
