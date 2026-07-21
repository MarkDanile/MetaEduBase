#!/usr/bin/env python3
"""Generate synthetic, review-required park cooperation notes.

Only stable ``客户ID`` relations and non-PII categories are read from the
customer workbook. Every generated fact is marked ``synthetic`` / ``待审核``
and must not be treated as a real customer interaction before review.
"""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook, load_workbook

_OUTPUT_NAME = "13_客户_合作跟进记录_待审核.xlsx"
_FIELDS = ("客户ID", "所属行业", "来源渠道")
_STAGES = ("初步接触", "需求确认", "方案沟通", "商务谈判", "签约准备")
_NEXT_DATES = (
    "2026-07-24",
    "2026-07-28",
    "2026-08-01",
    "2026-08-05",
    "2026-08-08",
)
_INTENT_BY_INDUSTRY = {
    "集成电路": "研发办公及中试空间",
    "生物医药": "研发实验及配套办公空间",
    "精密制造": "生产研发一体化空间",
    "新材料": "研发及小试空间",
    "人工智能": "研发办公空间",
    "其他": "办公及产业配套空间",
}
_NOTES = {
    "初步接触": "[模拟待审核] 建议核实企业扩租/入园计划、预算区间与决策链。",
    "需求确认": "[模拟待审核] 建议确认面积、层高、承重、能耗与交付时间要求。",
    "方案沟通": "[模拟待审核] 建议准备匹配房源、租赁条件与产业政策说明。",
    "商务谈判": "[模拟待审核] 建议核对报价口径、免租期、装修期与审批节点。",
    "签约准备": "[模拟待审核] 建议完成主体资质、合同条款与履约材料复核。",
}


def generate(source: Path, output: Path, limit: int) -> int:
    workbook = load_workbook(source, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    first_row = next(rows, None)
    if first_row is None:
        workbook.close()
        raise ValueError("source workbook is empty")
    headers = [str(value) for value in first_row]
    missing = [field for field in _FIELDS if field not in headers]
    if missing:
        workbook.close()
        raise ValueError(f"source workbook missing columns: {', '.join(missing)}")
    indexes = {field: headers.index(field) for field in _FIELDS}
    customers: list[dict[str, str]] = []
    for row in rows:
        item = {field: str(row[indexes[field]] or "").strip() for field in _FIELDS}
        if item["客户ID"]:
            customers.append(item)
        if len(customers) >= limit:
            break
    workbook.close()

    generated = Workbook()
    sheet = generated.active
    sheet.title = "合作跟进记录_待审核"
    sheet.append(
        [
            "跟进ID",
            "客户ID",
            "线索来源",
            "跟进阶段",
            "对接人角色",
            "跟进记录",
            "意向项目",
            "下次跟进时间",
            "数据来源",
            "审核状态",
            "生成规则版本",
            "备注",
        ]
    )
    for index, customer in enumerate(customers, 1):
        stage = _STAGES[(index - 1) % len(_STAGES)]
        sheet.append(
            [
                f"SYN-NOTE-{index:04d}",
                customer["客户ID"],
                customer["来源渠道"] or "待核实",
                stage,
                "企业联系人/决策人（待核实）",
                _NOTES[stage],
                _INTENT_BY_INDUSTRY.get(
                    customer["所属行业"], _INTENT_BY_INDUSTRY["其他"]
                ),
                _NEXT_DATES[(index - 1) % len(_NEXT_DATES)],
                "synthetic",
                "待审核",
                "park-followup-v1",
                "仅供开发测试；非真实客户沟通事实，审核通过前不得用于业务决策。",
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    generated.save(output)
    return len(customers)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("directory", type=Path)
    parser.add_argument("--limit", type=int, default=180)
    args = parser.parse_args()
    if args.limit < 1:
        raise SystemExit("--limit must be positive")
    source = args.directory / "05_客户.xlsx"
    output = args.directory / _OUTPUT_NAME
    if not source.is_file():
        raise SystemExit(f"source workbook not found: {source}")
    count = generate(source, output, args.limit)
    print(f"created={output} rows={count} status=synthetic_review_required")


if __name__ == "__main__":
    main()
